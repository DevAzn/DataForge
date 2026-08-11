"""Infer schema tree + history samples from uploaded text files."""
from __future__ import annotations

import csv
import io
import json
import re
import uuid
from typing import Any

import yaml

try:
    import xml.etree.ElementTree as ET
except ImportError:  # pragma: no cover
    ET = None  # type: ignore


def new_id() -> str:
    return str(uuid.uuid4())


def detect_format(file_name: str, text: str) -> str:
    lower = file_name.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return "xlsx"
    if lower.endswith(".json") or lower.endswith(".jsonl") or lower.endswith(".ndjson"):
        return "json"
    if lower.endswith((".yml", ".yaml")):
        return "yaml"
    if lower.endswith(".xml"):
        return "xml"
    if lower.endswith(".csv"):
        return "csv"
    t = text.strip()
    if t.startswith("<"):
        return "xml"
    if t.startswith("{") or t.startswith("["):
        return "json"
    if "," in t and "\n" in t:
        return "csv"
    return "txt"


def infer_schema_from_xlsx(file_name: str, raw: bytes) -> dict:
    """
    Infer a flat tabular schema from the first non-empty sheet of an .xlsx workbook.
    Header row → value fields with sample cells from the first data row.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = None
        for sheet in wb.worksheets:
            # pick first sheet that has any cell
            for row in sheet.iter_rows(max_row=5, values_only=True):
                if any(c is not None and str(c).strip() != "" for c in row):
                    ws = sheet
                    break
            if ws is not None:
                break
        if ws is None:
            ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ValueError("Empty workbook")
        keys: list[str] = []
        for i, h in enumerate(header):
            label = str(h).strip() if h is not None else ""
            keys.append(label or f"col_{i + 1}")
        # de-dupe keys
        seen: dict[str, int] = {}
        uniq: list[str] = []
        for k in keys:
            n = seen.get(k, 0)
            seen[k] = n + 1
            uniq.append(k if n == 0 else f"{k}_{n + 1}")
        keys = uniq
        data_rows: list[tuple] = []
        for row in rows_iter:
            if row is None:
                continue
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            data_rows.append(row)
            if len(data_rows) >= 20:
                break
        first = data_rows[0] if data_rows else tuple()
        root_rows = []
        for i, key in enumerate(keys):
            sample = first[i] if i < len(first) else None
            root_rows.append(
                {
                    "id": new_id(),
                    "key": key,
                    "kind": "value",
                    "sampleValue": "" if sample is None else str(sample),
                    "isPrimary": False,
                    "isUnique": False,
                    "children": [],
                    "sortOrder": i,
                }
            )
        if not root_rows:
            root_rows = [
                {
                    "id": new_id(),
                    "key": "field",
                    "kind": "value",
                    "sampleValue": "",
                    "isPrimary": False,
                    "isUnique": False,
                    "children": [],
                    "sortOrder": 0,
                }
            ]
        base = re.sub(r"\.[^.\\/]+$", "", file_name.replace("\\", "/").split("/")[-1]) or "Imported"
        history = []
        for row in data_rows[:5]:
            for i, key in enumerate(keys):
                if i < len(row) and row[i] is not None:
                    history.append(
                        {"categoryName": key, "keyName": key, "value": str(row[i])}
                    )
        return {
            "schema": {
                "id": new_id(),
                "name": base,
                "root": root_rows,
                "sourceFileName": file_name.replace("\\", "/").split("/")[-1],
                "sourceFormat": "xlsx",
            },
            "format": "xlsx",
            "recordHint": max(1, len(data_rows)),
            "scannedRecords": len(data_rows),
            "historySamples": history,
        }
    finally:
        wb.close()

def _sample_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (dict, list)):
        return None
    return str(v)


def _kind(v: Any) -> str:
    if isinstance(v, list):
        return "array"
    if isinstance(v, dict):
        return "object"
    return "value"


def infer_row(key: str, value: Any, sort: int = 0) -> dict:
    kind = _kind(value)
    if kind == "array":
        first = next((x for x in value if isinstance(x, dict)), value[0] if value else None)
        if isinstance(first, dict):
            children = [
                infer_row(k, first[k], i) for i, k in enumerate(first.keys())
            ]
        else:
            children = [
                {
                    "id": new_id(),
                    "key": "item",
                    "kind": "value",
                    "sampleValue": _sample_str(first) or "",
                    "isPrimary": False,
                    "isUnique": False,
                    "children": [],
                    "sortOrder": 0,
                }
            ]
        return {
            "id": new_id(),
            "key": key,
            "kind": "array",
            "isPrimary": False,
            "isUnique": False,
            "children": children,
            "sortOrder": sort,
        }
    if kind == "object":
        children = [
            infer_row(k, value[k], i) for i, k in enumerate(value.keys())
        ]
        return {
            "id": new_id(),
            "key": key,
            "kind": "object",
            "isPrimary": False,
            "isUnique": False,
            "children": children,
            "sortOrder": sort,
        }
    return {
        "id": new_id(),
        "key": key,
        "kind": "value",
        "sampleValue": _sample_str(value) or "",
        "isPrimary": False,
        "isUnique": False,
        "children": [],
        "sortOrder": sort,
    }


def merge_shapes(objects: list[dict], max_n: int = 100) -> dict:
    out: dict = {}
    for obj in objects[:max_n]:
        if not isinstance(obj, dict):
            continue
        for k, v in obj.items():
            if k not in out:
                out[k] = v
            else:
                # prefer dict over scalar for nesting
                if isinstance(v, dict) and not isinstance(out[k], dict):
                    out[k] = v
                elif isinstance(v, list) and not isinstance(out[k], list):
                    out[k] = v
    return out


def xml_to_dict(elem) -> Any:
    children = list(elem)
    if not children:
        return (elem.text or "").strip()
    d: dict[str, Any] = {}
    for child in children:
        tag = child.tag.split("}")[-1]
        val = xml_to_dict(child)
        if tag in d:
            if not isinstance(d[tag], list):
                d[tag] = [d[tag]]
            d[tag].append(val)
        else:
            d[tag] = val
    return d


def parse_csv(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def harvest_history(records: list[Any], max_records: int = 500, max_per_key: int = 200) -> list[dict]:
    bags: dict[str, list[str]] = {}

    def walk(obj: Any, path: list[str]) -> None:
        if isinstance(obj, dict):
            for k, v in obj.items():
                walk(v, path + [str(k)])
        elif isinstance(obj, list):
            for item in obj[:5]:
                walk(item, path)
        else:
            if obj is None:
                return
            key = ".".join(path) if path else "value"
            bags.setdefault(key, [])
            s = str(obj)
            if s and s not in bags[key] and len(bags[key]) < max_per_key:
                bags[key].append(s)

    for rec in records[:max_records]:
        walk(rec, [])

    out: list[dict] = []
    for key, vals in bags.items():
        for v in vals:
            out.append({"categoryName": key, "keyName": key, "value": v})
    return out


def infer_schema_from_file(file_name: str, content: str) -> dict:
    fmt = detect_format(file_name, content)
    text = content
    data: Any
    records: list[Any]

    if fmt == "json":
        t = text.strip()
        if "\n" in t and not t.startswith("[") and not t.startswith("{"):
            # ndjson
            lines = [json.loads(line) for line in t.splitlines() if line.strip()]
            data = lines
        else:
            data = json.loads(t)
    elif fmt == "yaml":
        data = yaml.safe_load(text)
    elif fmt == "xml":
        if ET is None:
            raise RuntimeError("XML parser unavailable")
        root = ET.fromstring(text)
        data = {root.tag.split("}")[-1]: xml_to_dict(root)}
    elif fmt == "csv":
        data = parse_csv(text)
    else:
        t = text.strip()
        if t.startswith("{") or t.startswith("["):
            data = json.loads(t)
            fmt = "json"
        elif "," in t and "\n" in t:
            data = parse_csv(t)
            fmt = "csv"
        else:
            data = {"content": text}

    if isinstance(data, list):
        records = data
        shape_src = [r for r in data if isinstance(r, dict)]
        shape = merge_shapes(shape_src) if shape_src else {}
        root_rows = [infer_row(k, shape[k], i) for i, k in enumerate(shape.keys())]
        if not root_rows and records:
            root_rows = [infer_row("item", records[0], 0)]
    elif isinstance(data, dict):
        records = [data]
        root_rows = [infer_row(k, data[k], i) for i, k in enumerate(data.keys())]
    else:
        records = [{"value": data}]
        root_rows = [
            {
                "id": new_id(),
                "key": "value",
                "kind": "value",
                "sampleValue": str(data),
                "isPrimary": False,
                "isUnique": False,
                "children": [],
                "sortOrder": 0,
            }
        ]

    if not root_rows:
        root_rows = [
            {
                "id": new_id(),
                "key": "field",
                "kind": "value",
                "sampleValue": "",
                "isPrimary": False,
                "isUnique": False,
                "children": [],
                "sortOrder": 0,
            }
        ]

    base = re.sub(r"\.[^.\\/]+$", "", file_name.replace("\\", "/").split("/")[-1]) or "Imported"
    history = harvest_history(records)
    return {
        "schema": {
            "id": new_id(),
            "name": base,
            "root": root_rows,
            "sourceFileName": file_name.replace("\\", "/").split("/")[-1],
            "sourceFormat": fmt,
        },
        "format": fmt,
        "recordHint": len(records),
        "scannedRecords": min(len(records), 500),
        "historySamples": history,
    }
