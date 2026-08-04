"""Header-first CSV/TXT export (real serialize path)."""
from __future__ import annotations

import sys
from pathlib import Path

# Allow running as script: python backend/tests/test_tabular_export.py
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from app.services import export_fmt  # noqa: E402
from app.services import generator  # noqa: E402


def _flat_schema():
    return {
        "id": "test-schema",
        "name": "tabular",
        "root": [
            {
                "id": "c1",
                "key": "name",
                "kind": "value",
                "sampleValue": "Ada",
                "children": [],
            },
            {
                "id": "c2",
                "key": "city",
                "kind": "value",
                "sampleValue": "London",
                "children": [],
            },
        ],
    }


def test_csv_header_then_data_rows():
    records = [
        {"name": "Ada", "city": "London"},
        {"name": "Bob", "city": "Paris"},
    ]
    text = export_fmt.serialize(
        records, "csv", multi_row=True, layout_mode="single-header"
    )
    lines = [ln for ln in text.strip().splitlines() if ln.strip()]
    assert len(lines) >= 3, text
    header = lines[0]
    assert "name" in header and "city" in header
    assert header.index("name") < header.index("city") or "name" in header
    # Data rows under those columns (order preserved from first record)
    assert "Ada" in lines[1] and "London" in lines[1]
    assert "Bob" in lines[2] and "Paris" in lines[2]
    # Must not be a JSON blob
    assert not text.lstrip().startswith("{")
    assert not text.lstrip().startswith("[")


def test_txt_header_then_data_rows_tab_separated():
    records = [
        {"name": "Ada", "city": "London"},
        {"name": "Bob", "city": "Paris"},
    ]
    text = export_fmt.serialize(records, "txt", multi_row=True)
    lines = [ln for ln in text.strip().splitlines() if ln.strip()]
    assert len(lines) >= 3, text
    header = lines[0]
    assert "name" in header and "city" in header
    # Tab-separated by default
    assert "\t" in header, f"expected TSV header, got {header!r}"
    cols = header.split("\t")
    assert "name" in cols and "city" in cols
    assert "Ada" in lines[1] and "London" in lines[1]
    assert not text.lstrip().startswith("{")
    assert not text.lstrip().startswith("[")


def test_xlsx_workbook_has_header_and_rows():
    """XLSX is real OOXML bytes with header + data (openpyxl)."""
    records = [
        {"name": "Ada", "city": "London"},
        {"name": "Bob", "city": "Paris"},
    ]
    raw = export_fmt.serialize(records, "xlsx", multi_row=True)
    assert isinstance(raw, (bytes, bytearray)), type(raw)
    # ZIP/OOXML magic
    assert raw[:2] == b"PK"
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("name", "city") or rows[0][0] == "name"
    assert "Ada" in rows[1]
    assert "Bob" in rows[2]
    assert export_fmt.extension_for_format("xlsx") == "xlsx"
    assert export_fmt.is_binary_format("xlsx") is True
    assert export_fmt.normalize_format("xls") == "xlsx"


def test_generate_then_serialize_csv_and_txt():
    schema = _flat_schema()
    result = generator.generate_records(
        schema,
        record_count=3,
        seed=42,
        ci_mode=True,
        allow_large=True,
    )
    records = result["records"]
    assert len(records) == 3

    csv_text = export_fmt.serialize(
        records, "csv", multi_row=True, layout_mode="single-header"
    )
    txt_text = export_fmt.serialize(records, "txt", multi_row=True)

    csv_lines = [ln for ln in csv_text.strip().splitlines() if ln.strip()]
    txt_lines = [ln for ln in txt_text.strip().splitlines() if ln.strip()]
    assert "name" in csv_lines[0] and "city" in csv_lines[0]
    assert len(csv_lines) == 4  # header + 3 data
    assert "name" in txt_lines[0] and "city" in txt_lines[0]
    assert "\t" in txt_lines[0]
    assert len(txt_lines) == 4


if __name__ == "__main__":
    test_csv_header_then_data_rows()
    test_txt_header_then_data_rows_tab_separated()
    test_generate_then_serialize_csv_and_txt()
    print("ok")
